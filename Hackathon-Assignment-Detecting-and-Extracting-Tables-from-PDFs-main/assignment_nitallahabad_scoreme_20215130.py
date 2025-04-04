# Author: Manish Kumar
# Project: Streamlit PDF Table Extractor App (Without Tabula or Camelot)

import io
import os
import re
import base64
import numpy as np
import pandas as pd
import pdfplumber
from PyPDF2 import PdfReader
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import streamlit as st

# Main function to extract tables from PDF

def process_pdf_tables(pdf_file, output_excel_path=None):
    print("[INFO] Processing uploaded PDF...")
    if output_excel_path:
        wb = Workbook()
        ws = wb.active
        ws.title = "Extracted Tables"

    extracted_dataframes = []

    try:
        with pdfplumber.open(pdf_file) as pdf:
            combined_tables = []

            for i, page in enumerate(pdf.pages):
                print(f"[INFO] Page {i+1}/{len(pdf.pages)}")
                tables = page.extract_tables()

                # Try built-in extraction
                if tables:
                    print(f"[+] Found {len(tables)} tables (built-in)")
                    combined_tables.extend(tables)
                else:
                    print("[-] No built-in tables found, trying custom parsing")
                    # Try bank format parser
                    parsed = parse_bank_statement_page(page)
                    if parsed:
                        combined_tables.append(parsed)
                        print("[+] Table found using bank statement parser")
                    else:
                        # Fallback to spacing-based parser
                        fallback = infer_columns_from_text(page)
                        if fallback:
                            combined_tables.append(fallback)
                            print("[+] Table found using text-based parser")

            # Write to Excel if any table is found
            if combined_tables:
                sheet_idx = 1
                for t_index, raw_table in enumerate(combined_tables):
                    if raw_table:
                        df = pd.DataFrame(raw_table)
                        df = sanitize_dataframe(df)
                        extracted_dataframes.append(df)

                        if output_excel_path:
                            if t_index > 0:
                                ws = wb.create_sheet(f"Table {sheet_idx}")
                                sheet_idx += 1

                            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                                for c_idx, value in enumerate(row, 1):
                                    ws.cell(row=r_idx, column=c_idx, value=value)

                if output_excel_path:
                    wb.save(output_excel_path)
                    print(f"[✓] Saved output to {output_excel_path}")
                return extracted_dataframes, True
            else:
                print("[!] No tables extracted.")
                return [], False

    except Exception as e:
        print(f"[ERROR] {e}")
        return [], False

# Function to extract tables formatted like bank statements

def parse_bank_statement_page(page):
    text = page.extract_text()
    if not text:
        return None

    lines = text.split('\n')
    date_regex = re.compile(r'\d{2}-[A-Za-z]{3}-\d{4}')
    patterns = [
        re.compile(r'(\d{2}-[A-Za-z]{3}-\d{4})\s+(.*?)(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)\s+(\d{1,3}(?:,\d{3})*(?:\.\d{2})?(?:Dr|Cr)?)'),
        re.compile(r'(\d{2}-[A-Za-z]{3}-\d{4})\s+(.*?)(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)'),
        re.compile(r'(\d{2}-[A-Za-z]{3}-\d{4})\s+([A-Za-z].*)')
    ]
    meta = extract_account_info(lines)

    table = [["Date", "Description", "Debit", "Credit", "Balance"]]

    for line in lines:
        if any(x in line for x in ["BANK NAME", "Page No", "REPORT", "----"]):
            continue
        for pat in patterns:
            match = pat.search(line)
            if match and date_regex.search(line):
                date = match.group(1)
                desc = match.group(2).strip()
                amount = match.group(3).strip() if len(match.groups()) >= 3 else ""
                credit, debit = (amount, "") if any(x in desc for x in ["Cr", "Credit"]) else ("", amount)
                balance = match.group(4) if len(match.groups()) >= 4 else ""
                table.append([date, desc, debit, credit, balance])
                break

    if meta:
        for k, v in meta.items():
            table.insert(1, ["", k, v, "", ""])

    return table if len(table) > 1 else None

# Fallback method: estimate columns based on whitespace spacing

def infer_columns_from_text(page):
    text = page.extract_text()
    if not text:
        return None

    lines = text.split('\n')
    columns = []
    data = []

    # Estimate column positions
    for line in lines[:10]:
        pos = [m.start() for m in re.finditer(r'\S+', line)]
        if len(pos) > 3:
            columns.append(pos)

    if not columns:
        return None

    avg_pos = []
    max_len = max(len(pos) for pos in columns)
    for i in range(max_len):
        col = [p[i] for p in columns if i < len(p)]
        if col:
            avg_pos.append(sum(col) // len(col))

    # Extract rows based on average positions
    for line in lines:
        if not line.strip():
            continue
        row = []
        last = 0
        for p in avg_pos:
            if p > len(line):
                row.append("")
                continue
            row.append(line[last:p].strip())
            last = p
        if last < len(line):
            row.append(line[last:].strip())
        data.append(row)

    return data if data else None

# Extract metadata from bank header lines

def extract_account_info(lines):
    info = {}
    patterns = {
        'Account Number': re.compile(r'Account\s+No\s*:\s*([0-9]+)'),
        'Account Name': re.compile(r'A/C\s+Name\s*:\s*(.+)'),
        'Account Holder': re.compile(r'A/C\s+Holder\s*:\s*(.+)'),
        'Open Date': re.compile(r'Open\s+Date\s*:\s*(.+)'),
        'Interest Rate': re.compile(r'Interest\s+Rate\s*:\s*(.+)'),
        'Statement Period': re.compile(r'Statement\s+of\s+account\s+for\s+the\s+period\s+of\s+(.+)')
    }
    for line in lines:
        for key, pat in patterns.items():
            match = pat.search(line)
            if match:
                info[key] = match.group(1).strip()
    return info

# Clean and prepare the dataframe

def sanitize_dataframe(df):
    df = df.replace(['', None], np.nan).dropna(how='all').fillna('')
    if df.shape[0] > 0 and all(str(x).isupper() for x in df.iloc[0].dropna()):
        df.columns = df.iloc[0]
        df = df.drop(0)
    return df.reset_index(drop=True)

# Generates an Excel download link

def create_excel_download_link(df, filename, index=False):
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=index, sheet_name='Table')
    b64 = base64.b64encode(buffer.getvalue()).decode()
    return f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}">Download Excel file</a>'

# Streamlit UI entry point

def main():
    st.set_page_config(page_title="PDF Table Extractor", page_icon="📊", layout="wide")
    st.title("PDF Table Extractor")
    st.write("Upload a PDF file to extract tables (no Tabula/Camelot used)")

    st.sidebar.header("Options")
    st.sidebar.radio("Extraction Method", ["Automatic", "Built-in", "Bank Statement", "Text Table"])

    uploaded_file = st.file_uploader("Choose a PDF file", type="pdf")
    if uploaded_file:
        with open("temp.pdf", "wb") as f:
            f.write(uploaded_file.getbuffer())

        if st.button("Extract Tables"):
            with st.spinner("Extracting tables..."):
                dfs, success = process_pdf_tables("temp.pdf")
                if success:
                    st.success(f"Extracted {len(dfs)} tables")
                    tabs = st.tabs([f"Table {i+1}" for i in range(len(dfs))])
                    for i, (tab, df) in enumerate(zip(tabs, dfs)):
                        with tab:
                            st.dataframe(df)
                            st.markdown(create_excel_download_link(df, f"table_{i+1}.xlsx"), unsafe_allow_html=True)

                    if len(dfs) > 1:
                        all_xlsx = io.BytesIO()
                        with pd.ExcelWriter(all_xlsx, engine='openpyxl') as writer:
                            for i, df in enumerate(dfs):
                                df.to_excel(writer, sheet_name=f"Table {i+1}", index=False)
                        b64 = base64.b64encode(all_xlsx.getvalue()).decode()
                        st.markdown(f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="all_tables.xlsx">Download All Tables</a>', unsafe_allow_html=True)
                else:
                    st.error("No tables were detected")

            if os.path.exists("temp.pdf"):
                os.remove("temp.pdf")

if __name__ == "__main__":
    main()